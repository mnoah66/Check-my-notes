
from collections import defaultdict
import openpyxl
from openpyxl import load_workbook
from datetime import datetime, time, timedelta, date
my_file = openpyxl.load_workbook('Book2.xlsx', read_only=True)
ws = my_file.active

people = {}
for row in ws.iter_rows(row_offset=1):
  a = row[0] # Name
  b = row[1] # Date
  c = row[2] # Start
  d = row[3] # End
  if a.value:
    print(type(c.value)) # <class 'datetime.time'
    t = datetime.combine(date.min, c.value) - datetime.min
    print(t.total_seconds())
