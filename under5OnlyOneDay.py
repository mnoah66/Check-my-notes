

'''
people = {}
for row in ws.iter_rows(row_offset=1):
  a = row[0] # Name
  b = row[1] # Date
  c = row[2] # Start
  d = row[3] # End
  if a.value:
    # If name already exists
    if a.value in people.keys():
      # To do: if the location already exists

      for k, v in people.items():
        for x in v:
          if x == int(b.value):
            print('Location found!')
            print(v)
            people[a.value][b.value].append([c.value, d.value])

            #print('Location found!')
          else:
            #to-do: add new location
            print(b.value)
    
    # Else, add new name
    else:
      print('add name to dict')
      people[a.value] = {b.value:[[c.value,d.value]]}

print(people)
'''


def overlap(list_item, key, k):
  import pandas
  intervals = list_item
  overlapping = [ [s,e] for s in intervals for e in intervals if s is not e and s[1]>e[0] and s[0]<e[0] ]
  duplicate = [ [s,e] for s in intervals for e in intervals if s is not e and s[0]==e[0] ]
  unique_duplicate = [list(x) for x in set(tuple(x) for x in duplicate)]
  
  for x in overlapping:
    print("{0} has overlapping notes on {1}".format(key, k))
    
  for x in unique_duplicate:
    print("{0} has a duplcate note on {1}".format(key, k))

def overlapping_notes(ws):
  from collections import defaultdict
  people = defaultdict(dict)
  for row in ws.iter_rows(row_offset=1):
      p=row[1]
      l=row[2]
      s=row[4]
      e=row[5]
      if p.value:
        if p.value not in people:
            people[p.value] = defaultdict(list)
        people[p.value][l.value].append((s.value, e.value))
  for key, val in people.items():
    for k, v in val.items():
      overlap(v, key, k)

if __name__ == "__main__":
  from collections import defaultdict
  import openpyxl
  from openpyxl import load_workbook

  trngfile = openpyxl.load_workbook('Book1.xlsx', read_only=True)
  ws = trngfile.active
  overlapping_notes(ws)

  ##for k, v in people.items():
  #for item in v:
  #  for k, v in item.items():
  #    overlap(v)

  




'''

import pandas as pd

# define dataframe, or df = pd.read_excel('file.xlsx')
df = pd.read_excel('Book2.xlsx')
#df = pd.DataFrame({'Name': ['John']*3 + ['Jane']*2,
#                   'Location': [20, 20, 21, 20, 21],
#                   'Start': [2.00, 3.00, 2.00, 9.00, 2.00],
#                   'End': [4.00, 5.00, 4.00, 10.00, 4.00]})

# convert cols to integers
int_cols = ['Start', 'End']
df[int_cols] = df[int_cols].apply(pd.to_numeric, downcast='integer')

# define inner dictionary grouper and split into list of dictionaries
def loc_list(x):
    d = {loc: w[int_cols].values.tolist() for loc, w in x.groupby('Location')}
    return [{i: j} for i, j in d.items()]

# define outer dictionary grouper
people = {k: loc_list(v) for k, v in df.groupby('Name')}

for k, v in people.items():
  for item in v:
    for k, v in item.items():
      overlap(v)


'''
