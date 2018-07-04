import tkinter as tk
from tkinter.font import Font
from tkinter import filedialog,messagebox
import openpyxl
from openpyxl import load_workbook
from datetime import timedelta, date, time
import datetime
import csv
import io
import os 
from tkinter import ttk
import xlsxwriter
from checknotesfunctions import flaggedWords, oddDuration, oddTimes, underUnits, shortNote


class SampleApp(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)

        self.title("Check My Notes")

        # -------------------- MAIN GUI -----------------------------

        self.labelKeywords = tk.Label(self, text="FLAG THE FOLLOWING WORDS (Separate with comma)",font=Font(family='Arial', size=11))
        self.labelKeywords.pack(fill=tk.X,padx=50,pady=5)
        self.entryKeywords = tk.Text(self,font=Font(family='Arial', size=10),height=2)
        self.entryKeywords.pack(fill=tk.X,padx=50, pady=2)

     
        self.labelDurationsGreater = tk.Label(self, text="FLAG TOTAL MINUTES GREATER OR EQUAL TO",font=Font(family='Arial', size=11))
        self.labelDurationsGreater.pack(fill=tk.X, padx=50)
        self.spinDurationsGreater = tk.Entry(self, width=5,  font=Font(family='Helvetica', size=12))
        self.spinDurationsGreater.pack(padx=5,pady=2)
        
        
        
        self.labelDurationsLess = tk.Label(self, text="FLAG TOTAL MINUTES LESS OR EQUAL TO",font=Font(family='Arial', size=11))
        self.labelDurationsLess.pack(fill=tk.X, padx=50)
        self.spinDurationsLess = tk.Entry(self, width=5,  font=Font(family='Helvetica', size=12))
        self.spinDurationsLess.pack(padx=5,pady=2)

        
        self.labelNoteLength = tk.Label(self, text="FLAG NOTE LENGTH (CHARACTERS) LESS THAN",font=Font(family='Arial', size=11))
        self.labelNoteLength.pack(fill=tk.X, padx=50)
        self.spinNoteLength = tk.Entry(self, width=5,  font=Font(family='Helvetica', size=12))
        self.spinNoteLength.pack(padx=5,pady=2)


        self.labelStartAfter = tk.Label(self, text="FLAG NOTE WITH START TIME AFTER",font=Font(family='Arial', size=11))
        self.labelStartAfter.pack(fill=tk.X,padx=50)
        self.spinHourAfter = tk.Spinbox(self, values=("","01","02","03","04","05","06","07","08","09","10","11","12"),font=Font(family='Helvetica', size=12), width=5,readonlybackground='white')
        self.spinHourAfter.pack(padx=5)
        self.spinMinAfter = tk.Spinbox(self, values=("","01",
        "02",
        "03",
        "04",
        "05",
        "06",
        "07",
        "08",
        "09",
        "10",
        "11",
        "12",
        "13",
        "14",
        "15",
        "16",
        "17",
        "18",
        "19",
        "20",
        "21",
        "22",
        "23",
        "24",
        "25",
        "26",
        "27",
        "28",
        "29",
        "30",
        "31",
        "32",
        "33",
        "34",
        "35",
        "36",
        "37",
        "38",
        "39",
        "40",
        "41",
        "42",
        "43",
        "44",
        "45",
        "46",
        "47",
        "48",
        "49",
        "50",
        "51",
        "52",
        "53",
        "54",
        "55",
        "56",
        "57",
        "58",
        "59",
        ), font=Font(family='Helvetica', size=12), width=5,readonlybackground='white')
        self.spinMinAfter.pack(padx=5)
        self.spinAMPMafter = tk.Spinbox(self, values=("","AM","PM"), font=Font(family='Helvetica', size=12), width=5)
        self.spinAMPMafter.pack(padx=5)

        self.labelStartBefore = tk.Label(self, text="FLAG NOTE WITH START TIME BEFORE",font=Font(family='Arial', size=11))
        self.labelStartBefore.pack(fill=tk.X,padx=50)
        self.spinHourBefore = tk.Spinbox(self, values=("","01","02","03","04","05","06","07","08","09","10","11","12"), font=Font(family='Helvetica', size=12), width=5)
        self.spinHourBefore.pack(padx=5)
        self.spinMinBefore = tk.Spinbox(self, values=("","01",
        "02",
        "03",
        "04",
        "05",
        "06",
        "07",
        "08",
        "09",
        "10",
        "11",
        "12",
        "13",
        "14",
        "15",
        "16",
        "17",
        "18",
        "19",
        "20",
        "21",
        "22",
        "23",
        "24",
        "25",
        "26",
        "27",
        "28",
        "29",
        "30",
        "31",
        "32",
        "33",
        "34",
        "35",
        "36",
        "37",
        "38",
        "39",
        "40",
        "41",
        "42",
        "43",
        "44",
        "45",
        "46",
        "47",
        "48",
        "49",
        "50",
        "51",
        "52",
        "53",
        "54",
        "55",
        "56",
        "57",
        "58",
        "59",
        ), font=Font(family='Helvetica', size=12), width=5)
        self.spinMinBefore.pack(padx=5)
        self.spinAMPMbefore = tk.Spinbox(self, values=("","AM","PM"), font=Font(family='Helvetica', size=12), width=5)
        self.spinAMPMbefore.pack(padx=5)


        self.labelUnderUnits = tk.Label(self, text="FLAG INDIVIDUALS WITH TOTAL UNITS LESS THAN:",font=Font(family='Arial', size=11))
        self.labelUnderUnits.pack(fill=tk.X, padx=50)
        self.spinUnderUnits = tk.Entry(self, width=5,font=Font(family='Helvetica', size=12))
        self.spinUnderUnits.pack(padx=5,pady=2)
 
        ## Choose file to read
        self.buttonFile = tk.Button(self, text="1. Choose File to be scanned", command=self.file_choose)
        self.buttonFile.pack(fill=tk.X,padx=50,pady=2)
        self.labelFile = tk.Label(self, text="")
        self.labelFile.pack() 

        # Choose direcgtory to save file ot
        self.buttonFileOutput = tk.Button(self, text="2. Confirm output file location", command=self.folder_choose)
        self.buttonFileOutput.pack(fill=tk.X,padx=50,pady=2)
        self.labelFileOutput = tk.Entry(self, text="",background='grey94')
        self.labelFileOutput.pack(fill=tk.X,padx=50,pady=2) 
        
        self.button = tk.Button(self, text="3. RUN", command=self.on_button)
        self.button.pack(fill=tk.X,padx=50,pady=2) 
        

        self.labelWorking = tk.Label(self, text="")
        self.labelWorking.pack()

        import configparser
        global config 
        config = configparser.ConfigParser()
        config.read('config.ini')
        self.entryKeywords.insert(1.0, config.get('DEFAULT', 'entryKeywords'))
        self.spinDurationsGreater.insert(0, config.get('DEFAULT', 'spinDurationsGreater'))
        self.spinDurationsLess.insert(0, config.get('DEFAULT', 'spinDurationsLess'))
        self.spinNoteLength.insert(0, config.get('DEFAULT', 'spinNoteLength'))
        self.spinHourAfter.insert(0, config.get('DEFAULT', 'spinHourAfter'))
        self.spinMinAfter.insert(0, config.get('DEFAULT', 'spinMinAfter'))
        self.spinAMPMafter.insert(0, config.get('DEFAULT', 'spinAMPMafter'))
        self.spinHourBefore.insert(0, config.get('DEFAULT', 'spinHourBefore'))
        self.spinMinBefore.insert(0, config.get('DEFAULT', 'spinMinBefore'))
        self.spinAMPMbefore.insert(0, config.get('DEFAULT', 'spinAMPMbefore'))
        self.spinUnderUnits.insert(0,config.get('DEFAULT','spinUnderUnits'))
        self.labelFileOutput.insert(0,config.get('DEFAULT','labelFileOutput'))
        
        
    #-------------------------  END - MAIN GUI ---------------------------------------------------

    def folder_choose(self):
        global dirname
        dirname = filedialog.askdirectory(parent=self, initialdir="/", title='Please select a directory')
        self.labelFileOutput.delete(0, 'end')
        self.labelFileOutput.insert(0,dirname)
        self.labelFileOutput.pack(fill=tk.X,padx=50)    

    def file_choose(self):
        global file_name
        file_name = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_name.endswith(".xlsx"):
            pass
        else:
            return tk.messagebox.showerror("Warning - File", "Please choose '.xlsx' files only.")
        
        self.labelFile.configure(text=file_name)
        self.labelFile.pack(fill=tk.X,padx=50)
        
    def excel_writer(self):
        global excel_file_name 
        excel_file_name = str(self.labelFileOutput.get()) + '\CheckMyNotes-Created-' + str(datetime.datetime.now().strftime('%Y%m%d%H%M%S'))+ '.xlsx'
        workbook = xlsxwriter.Workbook(excel_file_name)
        worksheet = workbook.add_worksheet()

        format3 = workbook.add_format({'num_format': 'mm/dd/yy'})
        format7 = workbook.add_format({'num_format': 'hh:mm AM/PM'})

        row = 0
        col = 0

        worksheet.write(row,col, 'Flag')
        worksheet.write(row,col+1, 'Individual')
        worksheet.write(row,col+2, 'Start time')
        worksheet.write(row,col+3, 'End Time')
        worksheet.write(row,col+4, 'Date')
        worksheet.write(row,col+5, 'Note/Excerpt')
        worksheet.write(row,col+6, 'Program')
        worksheet.write(row,col+7, 'Duration')
        worksheet.write(row,col+8, 'Note writer')
        worksheet.write(row,col+9, 'Audit Comments')

        from operator import itemgetter
        alpha_list = sorted(results_list, key=itemgetter(1)) # Sort nested list based on the 1th value (individuals name)

        for item in (alpha_list):
            row += 1
            worksheet.write(row, col, item[0])      # The flagged phrase or word
            worksheet.write(row, col +1, item[1])   # The individual
            worksheet.write(row, col +2, item[2], format7)   # The start time
            worksheet.write(row, col +3, item[3], format7)   # The end time
            worksheet.write(row, col +4, item[4], format3)   # The Date
            worksheet.write(row, col +5, item[5])   # The note/excerpt
            worksheet.write(row, col +6, item[6])   # The program name
            worksheet.write(row, col +7, item[7])   # The Duration
            worksheet.write(row, col +8, item[8])   # The Duration
        workbook.close()

    def on_button(self):
        global results_list
        results_list = []

        self.saveConfig()
        print(self.entryKeywords.get("1.0", 'end-1c'))
        try:
            file_name
            with open(file_name, "rb") as f:
                in_mem_file = io.BytesIO(f.read())
            trngfile = openpyxl.load_workbook(in_mem_file, read_only=True)
            ws = trngfile.active
        except:
            return tk.messagebox.showerror("Warning - File", "An error occurred with the file.  Please choose an .xlsx file only.")

        if self.labelFileOutput.get() == "":
            return  tk.messagebox.showerror("Warning - Directory", "Please choose a save location.") 
        
        # ---------- VARIABLES TO PASS TO FUNCTIONS -----------------
        keywords = self.entryKeywords.get("1.0", 'end-1c')
        my_list = keywords.split(",")
        my_list = [x.strip().lower() for x in my_list]
        try:
            greaterthan = int(self.spinDurationsGreater.get())
            lessthan = int(self.spinDurationsLess.get())
            notelength = int(self.spinNoteLength.get())
            unitThreshold = int(self.spinUnderUnits.get())
        except (TypeError, ValueError):
            return tk.messagebox.showerror("Warning - Integer", "Please enter whole numbers only (e.g. 360 or 12)")
        
        startTimeAfter = self.spinHourAfter.get() + ":" + self.spinMinAfter.get() + " " + self.spinAMPMafter.get()
        startTimeBefore = self.spinHourBefore.get() + ":" + self.spinMinBefore.get() + " " + self.spinAMPMbefore.get()
        
        # iF THE KEYWORDS IS BLANK, USER DOES NOT WANT TO 
        # FLAG WORDS
        if self.entryKeywords.get("1.0",'end-1c').lower() == '':
            pass
        else:
            flaggedWords(ws, my_list, results_list)
        oddDuration(ws, greaterthan, lessthan, results_list)
        shortNote(ws, notelength, results_list)
        oddTimes(ws,startTimeAfter,startTimeBefore, results_list)
        underUnits(ws,unitThreshold, results_list)
        self.labelWorking.configure(font=Font(family='Helvetica', size=12),text="FINISHED!")
        self.excel_writer()

        def callback(event):
            import os
            import webbrowser
            webbrowser.open_new(r"file://" + os.path.abspath(str(excel_file_name)))
            self.link.configure(text="")
        
        self.link = tk.Label(self, text="Click here for file", fg="blue", cursor="hand2")
        self.link.pack()
        self.link.bind("<Button-1>", callback)
    def saveConfig(self):
        config.set('DEFAULT', 'entryKeywords', self.entryKeywords.get("1.0",'end-1c'))
        config.set('DEFAULT', 'spinDurationsGreater', self.spinDurationsGreater.get())
        config.set('DEFAULT', 'spinDurationsLess', self.spinDurationsLess.get())
        config.set('DEFAULT', 'spinNoteLength', self.spinNoteLength.get())
        config.set('DEFAULT', 'spinHourAfter', self.spinHourAfter.get())
        config.set('DEFAULT', 'spinMinAfter', self.spinMinAfter.get())
        config.set('DEFAULT', 'spinAMPMafter', self.spinAMPMafter.get())
        config.set('DEFAULT', 'spinHourBefore', self.spinHourBefore.get())
        config.set('DEFAULT', 'spinMinBefore', self.spinMinBefore.get())
        config.set('DEFAULT', 'spinAMPMbefore', self.spinAMPMbefore.get())
        config.set('DEFAULT', 'spinUnderUnits', self.spinUnderUnits.get())
        config.set('DEFAULT', 'labelFileOutput', self.labelFileOutput.get())
        config.write(open('config.ini','w'))
        
app = SampleApp()
app.mainloop()